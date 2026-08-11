# -*- coding: utf-8 -*-
"""
Genera los datos del informe de ventas de los 3 almacenes de colchones.

Lee la hoja MOVIMIENTOS de los 3 Excel, se queda SOLO con las filas cuyo
TIPO MOVIMIENTO es VENTA (las de COMPRAS son abastecimiento, no venta) y
escribe dos cosas en esta misma carpeta:

  datos.js                  lo que lee index.html (el informe web)
  VENTAS_3_ALMACENES.xlsx   el consolidado para abrir en Excel, con la
                            columna ALMACÉN y una columna por almacén en
                            las hojas de resumen

Cómo actualizarlo cuando lleguen Excel nuevos:
  1. Dejar los 3 archivos en la carpeta de Descargas.
  2. Ajustar la lista ARCH de abajo si cambiaron de nombre.
  3. python colchones/generar_datos.py
  4. Recargar el informe en el navegador.

Necesita openpyxl:  pip install openpyxl
"""
import openpyxl, os, json, datetime, collections
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- de dónde salen los Excel -------------------------------------------------
D = os.path.join(os.environ.get('USERPROFILE', ''), 'Downloads')
ARCH = [
    ('CERRADORA',      'MOV. CERRADORA (1) (1).xlsx'),
    ('NUEVO AMANECER', 'MOV. NUEVO AMANECER (1).xlsx'),
    ('REY',            'MOV. REY (1).xlsx'),
]
DEST = os.path.dirname(os.path.abspath(__file__))
ALM = [a for a, _ in ARCH]

DIA = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']
MESN = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO',
        'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']


def txt(v):
    return '' if v is None else str(v).strip()


def nm(v):
    """Number de una celda. Aguanta '$ 1.250.000' escrito a mano."""
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('$', '').replace(' ', '').replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


# ============================ 1. LEER LOS EXCEL ==============================
V, sin_valor = [], 0
for almacen, arch in ARCH:
    ruta = os.path.join(D, arch)
    if not os.path.exists(ruta):
        raise SystemExit('No encuentro el archivo: ' + ruta)
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if 'MOVIMIENTOS' not in wb.sheetnames:
        raise SystemExit(arch + ' no tiene hoja MOVIMIENTOS')
    it = wb['MOVIMIENTOS'].iter_rows(values_only=True)
    ix = {c: i for i, c in enumerate([txt(c) for c in next(it)])}

    def g(r, k):
        i = ix.get(k)
        return r[i] if i is not None and i < len(r) else None

    for r in it:
        if r is None or all(c is None or c == '' for c in r):
            continue
        f = g(r, 'FECHA')
        if isinstance(f, datetime.datetime):
            fecha = f.date()
        elif isinstance(f, datetime.date):
            fecha = f
        else:
            try:
                fecha = datetime.date.fromisoformat(txt(f)[:10])
            except Exception:
                continue
        # La columna del Excel trae un espacio adelante: ' TIPO MOVIMIENTO'
        tipo = (txt(g(r, ' TIPO MOVIMIENTO')) or txt(g(r, 'TIPO MOVIMIENTO'))).upper()
        if tipo != 'VENTA':
            continue
        can = abs(nm(g(r, 'CAN')))          # en las ventas la cantidad va en negativo
        vu = nm(g(r, 'VALOR UNITARIO'))
        tot = nm(g(r, 'TOTAL PRODUCTO'))
        if tot == 0 and vu and can:
            tot = vu * can                  # el total venía vacío: se reconstruye
        elif tot == 0:
            sin_valor += 1
        V.append({'almacen': almacen, 'fecha': str(fecha),
                  'producto': txt(g(r, 'PRODUCTO')), 'categoria': txt(g(r, 'CATEGORIA')),
                  'can': can, 'vu': vu, 'total': tot,
                  'pago': txt(g(r, 'FORMA DE PAGO')), 'cliente': txt(g(r, 'NOMBRE DEL CLIENTE'))})
    wb.close()

if not V:
    raise SystemExit('No salió ninguna venta. ¿Cambió el nombre de las columnas?')

# ============================ 2. datos.js ====================================
# Se guarda con diccionarios y filas como arreglos: el archivo pesa la cuarta
# parte que un JSON con nombres repetidos 10.000 veces.
def dic(campo):
    orden, vistos = [], {}
    for f in V:
        v = f[campo]
        if v not in vistos:
            vistos[v] = len(orden)
            orden.append(v)
    return vistos, orden


iprod, prod = dic('producto')
icat, cat = dic('categoria')
ipag, pag = dic('pago')
icli, cli = dic('cliente')

filas = sorted(
    [[ALM.index(f['almacen']), f['fecha'], iprod[f['producto']], icat[f['categoria']],
      round(f['can'], 2), round(f['vu']), round(f['total']), ipag[f['pago']], icli[f['cliente']]]
     for f in V],
    key=lambda r: (r[1], r[0]))

paq = {'generado': datetime.date.today().isoformat(), 'almacenes': ALM,
       'productos': prod, 'categorias': cat, 'pagos': pag, 'clientes': cli, 'filas': filas}
p = os.path.join(DEST, 'datos.js')
with open(p, 'w', encoding='utf-8') as fh:
    fh.write('window.DATOS=' + json.dumps(paq, ensure_ascii=False, separators=(',', ':')) + ';\n')
print(f'datos.js                 {os.path.getsize(p)/1024:>6.0f} KB  {len(filas)} ventas')

# ============================ 3. el Excel ====================================
wb = openpyxl.Workbook()
cabF = Font(bold=True, color='FFFDF8', size=10)
cabR = PatternFill('solid', fgColor='2A1F16')


def cabecera(ws, cols, anchos):
    ws.append(cols)
    for i, c in enumerate(cols, 1):
        cl = ws.cell(row=1, column=i)
        cl.font, cl.fill = cabF, cabR
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = anchos[i - 1]
    ws.freeze_panes = 'A2'


def plata(ws, desde_col, hasta_col):
    for r in range(2, ws.max_row + 1):
        for c in range(desde_col, hasta_col + 1):
            ws.cell(row=r, column=c).number_format = '#,##0'


# Hoja 1: todas las ventas, con la columna ALMACÉN para filtrar
ws = wb.active
ws.title = 'VENTAS'
cabecera(ws, ['ALMACÉN', 'FECHA', 'AÑO', 'MES', 'DÍA SEMANA', 'PRODUCTO', 'CATEGORÍA',
              'CANTIDAD', 'VALOR UNITARIO', 'TOTAL', 'FORMA DE PAGO', 'CLIENTE'],
         [17, 11, 6, 11, 11, 34, 20, 9, 14, 15, 27, 32])
for f in V:
    d = datetime.date.fromisoformat(f['fecha'])
    ws.append([f['almacen'], d, d.year, MESN[d.month], DIA[d.weekday()], f['producto'],
               f['categoria'], f['can'], round(f['vu']), round(f['total']), f['pago'], f['cliente']])
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'
plata(ws, 9, 10)
ws.auto_filter.ref = f'A1:L{ws.max_row}'

# Hoja 2: mes a mes, una columna por almacén, con las unidades al lado
ws2 = wb.create_sheet('MES A MES')
cabecera(ws2, ['MES'] + [c for a in ALM for c in ('$ ' + a, 'UND ' + a)] + ['TOTAL', 'UND TOTAL'],
         [14] + [17, 11] * len(ALM) + [18, 11])
mm = {}
for f in V:
    k = f['fecha'][:7]
    o = mm.setdefault(k, {a: [0, 0] for a in ALM})
    o[f['almacen']][0] += f['total']
    o[f['almacen']][1] += f['can']
for k in sorted(mm):
    o = mm[k]
    ws2.append([k] + [v for a in ALM for v in (round(o[a][0]), round(o[a][1]))] +
               [round(sum(o[a][0] for a in ALM)), round(sum(o[a][1] for a in ALM))])
ult = ws2.max_row + 1
ws2.append(['TOTAL'] +
           [v for a in ALM for v in (round(sum(mm[k][a][0] for k in mm)),
                                     round(sum(mm[k][a][1] for k in mm)))] +
           [round(sum(mm[k][a][0] for k in mm for a in ALM)),
            round(sum(mm[k][a][1] for k in mm for a in ALM))])
plata(ws2, 2, ws2.max_column)
for c in range(1, ws2.max_column + 1):
    ws2.cell(row=ult, column=c).font = Font(bold=True, color='B5651D')

# Hoja 3: qué día de la semana se vende más (lo comparable es el promedio por día)
ws3 = wb.create_sheet('DÍA DE LA SEMANA')
cabecera(ws3, ['DÍA'] + [a + ' (prom/día)' for a in ALM] +
         ['TOTAL (prom/día)', 'TOTAL VENDIDO', 'DÍAS', 'VENTAS'],
         [12] + [20] * len(ALM) + [20, 18, 8, 9])
sem = {w: {'t': {a: 0 for a in ALM}, 'd': {a: set() for a in ALM},
           'dt': set(), 'n': 0, 'tt': 0} for w in range(7)}
for f in V:
    o = sem[datetime.date.fromisoformat(f['fecha']).weekday()]
    o['t'][f['almacen']] += f['total']
    o['d'][f['almacen']].add(f['fecha'])
    o['dt'].add(f['fecha'])
    o['n'] += 1
    o['tt'] += f['total']
for w in range(7):
    o = sem[w]
    ws3.append([DIA[w]] +
               [round(o['t'][a] / len(o['d'][a])) if o['d'][a] else 0 for a in ALM] +
               [round(o['tt'] / len(o['dt'])) if o['dt'] else 0,
                round(o['tt']), len(o['dt']), o['n']])
plata(ws3, 2, ws3.max_column)

# Hoja 4: precio de cada referencia. El precio promedio es PONDERADO
# (plata / unidades), no el promedio de los precios de lista.
ws4 = wb.create_sheet('PRECIOS Y PRODUCTOS')
cols4 = (['REFERENCIA'] + ['$ ' + a for a in ALM] + ['PRECIO PROM. ' + a for a in ALM] +
         ['PRECIO PROMEDIO', 'PRECIO MÁS BAJO', 'PRECIO MÁS ALTO',
          'UNIDADES', 'PLATA VENDIDA', 'LÍNEAS'])
cabecera(ws4, cols4, [34] + [16] * len(ALM) * 2 + [16, 15, 15, 10, 17, 9])
pr = {}
for f in V:
    o = pr.setdefault(f['producto'], {'t': {a: 0 for a in ALM}, 'q': {a: 0 for a in ALM},
                                      'tt': 0, 'qq': 0, 'n': 0, 'mn': None, 'mx': 0})
    o['t'][f['almacen']] += f['total']
    o['q'][f['almacen']] += f['can']
    o['tt'] += f['total']
    o['qq'] += f['can']
    o['n'] += 1
    precio = f['vu'] if f['vu'] > 0 else (f['total'] / f['can'] if f['can'] else 0)
    if precio > 0:
        o['mn'] = precio if o['mn'] is None else min(o['mn'], precio)
        o['mx'] = max(o['mx'], precio)
for nom, o in sorted(pr.items(), key=lambda kv: -kv[1]['tt']):
    ws4.append([nom] +
               [round(o['t'][a]) for a in ALM] +
               [round(o['t'][a] / o['q'][a]) if o['q'][a] else None for a in ALM] +
               [round(o['tt'] / o['qq']) if o['qq'] else None,
                round(o['mn']) if o['mn'] else None, round(o['mx']) or None,
                round(o['qq']), round(o['tt']), o['n']])
plata(ws4, 2, ws4.max_column)
ws4.auto_filter.ref = f'A1:{get_column_letter(len(cols4))}{ws4.max_row}'

# Hoja 5: precio por cliente y por referencia (a quién le venden más barato)
ws5 = wb.create_sheet('PRECIO POR CLIENTE')
cabecera(ws5, ['CLIENTE', 'REFERENCIA', 'UNIDADES', 'PRECIO QUE PAGA',
               'PRECIO MÁS BAJO', 'PRECIO MÁS ALTO', 'PLATA', 'LOS DEMÁS PAGAN', 'DIFERENCIA %'],
         [30, 32, 10, 16, 16, 16, 16, 17, 13])
cp = {}
for f in V:
    o = cp.setdefault((f['cliente'], f['producto']), {'t': 0, 'q': 0, 'mn': None, 'mx': 0})
    o['t'] += f['total']
    o['q'] += f['can']
    precio = f['vu'] if f['vu'] > 0 else (f['total'] / f['can'] if f['can'] else 0)
    if precio > 0:
        o['mn'] = precio if o['mn'] is None else min(o['mn'], precio)
        o['mx'] = max(o['mx'], precio)
for (clie, ref), o in sorted(cp.items(), key=lambda kv: (kv[0][0], -kv[1]['t'])):
    prom = o['t'] / o['q'] if o['q'] else 0
    # lo que paga TODO EL RESTO por la misma referencia
    rt = sum(x['t'] for (c2, r2), x in cp.items() if r2 == ref and c2 != clie)
    rq = sum(x['q'] for (c2, r2), x in cp.items() if r2 == ref and c2 != clie)
    otros = rt / rq if rq else 0
    ws5.append([clie, ref, round(o['q']), round(prom),
                round(o['mn']) if o['mn'] else None, round(o['mx']) or None, round(o['t']),
                round(otros) if otros else None,
                round((prom / otros - 1) * 100, 1) if otros and prom else None])
plata(ws5, 3, 8)
ws5.auto_filter.ref = f'A1:I{ws5.max_row}'

# Hoja 6: formas de pago
ws6 = wb.create_sheet('FORMA DE PAGO')
cabecera(ws6, ['FORMA DE PAGO'] + ALM + ['TOTAL'], [32] + [18] * len(ALM) + [18])
fp = {}
for f in V:
    fp.setdefault(f['pago'] or '(sin dato)', {a: 0 for a in ALM})[f['almacen']] += f['total']
for nom, o in sorted(fp.items(), key=lambda kv: -sum(kv[1].values())):
    ws6.append([nom] + [round(o[a]) for a in ALM] + [round(sum(o.values()))])
plata(ws6, 2, ws6.max_column)

px = os.path.join(DEST, 'VENTAS_3_ALMACENES.xlsx')
wb.save(px)
print(f'VENTAS_3_ALMACENES.xlsx  {os.path.getsize(px)/1024:>6.0f} KB  6 hojas')

tot = sum(f['total'] for f in V)
print(f'\n{len(V)} ventas · ${round(tot):,} · {len(pr)} referencias · {len(cli)} clientes')
for a in ALM:
    t = sum(f['total'] for f in V if f['almacen'] == a)
    print(f'  {a:16} ${round(t):>15,}   {t/tot*100:>5.1f}%')
if sin_valor:
    print(f'\nOJO: {sin_valor} líneas de venta quedaron en $0 porque el Excel '
          f'no traía ni valor unitario ni total.')
